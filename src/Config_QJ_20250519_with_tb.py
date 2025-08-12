from src.TrackCircuitElement.SectionGroup import *
from src.TrackCircuitElement.Train import *
from src.TrackCircuitElement.Line import *
from src.TrackCircuitElement.LineGroup import *
from src.Model.MainModel import *
from src.Model.ModelParameter import *
from src.FrequencyType import Freq
from src.Model.PreModel import PreModel
from src.logMethod import *

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

import re
import os
import time
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib import cm
plt.rcParams['font.sans-serif'] = ['SimSun']
plt.rcParams['mathtext.fontset'] = 'stix'
plt.rcParams['axes.unicode_minus'] = False


# 配置输入
def config_input_20250519_with_tb(list_type, list_length, list_freq):

    columns = [
        '序号',

        '主串类型',
        '被串类型',

        '区段长度',
        '主串频率',
        '被串频率',
        '相对位置',

        '主串分路位置',
        '是否换装TB',


        # '主串方向',
        # '被串方向',
    ]

    df = pd.DataFrame(index=columns, dtype='object')

    counter = 1
    for val_type in list_type:
        for val_length in list_length:
            for val_freq in list_freq:

                length = val_length

                offset = -length
                while offset <= length:
                # offset = 0
                # while offset <= 0:

                    # list_zhu_sht = list(range(0, length, 50))
                    list_zhu_sht = list()
                    list_zhu_sht.insert(0, '同位置')
                    list_zhu_sht.insert(0, '调整')
                    for val_zhu_sht in list_zhu_sht:

                        for val_tb in [False, True]:
                            s0 = pd.Series(name=counter, index=columns)

                            s0['序号'] = s0.name

                            s0['主串类型'] = val_type[0]
                            s0['被串类型'] = val_type[1]

                            s0['区段长度'] = length

                            s0['主串频率'] = val_freq[0]
                            s0['被串频率'] = val_freq[1]

                            s0['相对位置'] = offset

                            s0['主串分路位置'] = val_zhu_sht
                            s0['是否换装TB'] = val_tb

                            print('generate row: %s --> %s' % (counter, s0.tolist()))

                            df = pd.concat([df, s0], axis=1)
                            counter += 1

                            # if counter == 5:
                            #     df = df.transpose()
                            #     return df

                    offset += 100

    df = df.transpose()

    return df


# 配置表头
def config_headlist_20250519_with_tb():
    head_list = [
        '序号',
        # '备注',
        # '线路名称', '车站名称',
        # '主串区段', '被串区段',

        # '线间距(m)',
        '耦合系数(μH/km)',
        # '并行长度(m)',
        # '被串相对位置(m)',

        '主串区段类型', '被串区段类型',
        '主串方向', '被串方向',
        '主串分路位置(m)',
        '是否换装TB',

        # '线间距(m)',
        # '耦合系数(μH/km)',
        # '并行长度(m)',

        # '调谐区错位(m)',
        '被串相对位置(m)',
        '主串区段长度(m)', '被串区段长度(m)',
        # '主串左端坐标', '被串左端坐标',

        # '主串区段类型', '被串区段类型',
        '主串频率(Hz)', '被串频率(Hz)',

        '主串电容数(含TB)', '被串电容数(含TB)',
        '主串电容值(μF)', '被串电容值(μF)',

        '主串电容数量列表', '被串电容数量列表',
        '主串电容容值列表', '被串电容容值列表',

        '钢轨电阻(Ω/km)', '钢轨电感(H/km)',

        '主串道床电阻(Ω·km)', '被串道床电阻(Ω·km)',

        # '主串TB模式', '被串TB模式',

        '主串分路电阻(Ω)', '被串分路电阻(Ω)',
        '主串电缆长度(km)', '被串电缆长度(km)',

        '分路间隔(m)',

        '主串电平级',
        '电源电压',

        '被串最大干扰电流(A)', '被串最大干扰位置(m)',
        # '故障位置', '故障类型',
        # '干扰值变化',
    ]

    return head_list


def config_c_num_20250519_with_tb(freq: Freq, length, sec_type):
    freq_value = freq.value

    if 0 < length <= 300:
        key = 0
    elif length > 300:
        key = int((length - 251) / 50)
    else:
        raise KeyboardInterrupt('config_c_num error: 区段长度错误')

    table_j = {
        0: [0, 0, 0, 0],
        1: [4, 4, 3, 3],
        2: [4, 4, 4, 4],
        3: [5, 5, 5, 6],
        4: [6, 6, 6, 6],
        5: [6, 6, 6, 6],
        6: [6, 7, 6, 7],
        7: [7, 7, 7, 7],
        8: [8, 8, 8, 8],
        9: [8, 8, 8, 8],
        10: [9, 9, 9, 9],
        11: [9, 9, 9, 9],
        12: [9, 9, 9, 9],
        13: [10, 10, 10, 10],
        14: [10, 10, 10, 10],
        15: [11, 11, 11, 11],
        16: [11, 11, 11, 11],
        17: [12, 12, 12, 12],
        18: [12, 12, 12, 14],
        19: [13, 13, 13, 15],
        20: [14, 14, 16, 16],
        21: [16, 16, 18, 20],
        22: [18, 18, 18, 20],
    }

    table_t = {
        0: [0, 0, 0, 0],
        1: [5, 5, 4, 4],
        2: [6, 6, 5, 5],
        3: [7, 7, 5, 5],
        4: [8, 8, 6, 6],
        5: [9, 9, 7, 7],
        6: [10, 10, 7, 7],
        7: [10, 10, 8, 8],
        8: [11, 11, 8, 8],
        9: [12, 12, 9, 9],
        10: [13, 13, 10, 10],
        11: [14, 14, 10, 10],
        12: [15, 15, 11, 11],
        13: [15, 15, 12, 12],
        14: [16, 16, 12, 12],
        15: [17, 17, 13, 13],
        16: [18, 18, 13, 13],
        17: [19, 19, 14, 14],
        18: [20, 20, 15, 15],
        19: [20, 20, 15, 15],
        20: [21, 21, 16, 16],
        21: [22, 22, 17, 17],
        22: [23, 23, 17, 17],
    }

    if sec_type == '普速':
        table = table_j
    elif sec_type == '客专':
        table = table_t
    else:
        raise KeyboardInterrupt('config_c_num error: 区段类型错误')

    if key not in table.keys():
        raise KeyboardInterrupt('config_c_num error: 区段长度超长')

    index_dict = {
        1700: 0,
        2000: 1,
        2300: 2,
        2600: 3,
    }

    if freq_value not in index_dict.keys():
        raise KeyboardInterrupt('config_c_num error: 区段频率错误')

    c_num = table[key][index_dict[freq_value]]
    return c_num


def config_c_value_20250519_with_tb(freq: Freq, sec_type):
    freq_value = freq.value

    value_dict = {
        1700: 55,
        2000: 50,
        2300: 46,
        2600: 40,
    }

    if freq_value not in value_dict.keys():
        raise KeyboardInterrupt('config_c_value_imp error: 区段频率错误')

    if sec_type == '普速':
        c_value = value_dict[freq_value]
    elif sec_type == '客专':
        c_value = 25
    else:
        raise KeyboardInterrupt('config_c_value_imp error: 区段类型错误')

    return c_value


def config_c_pack_20250519_with_tb(freq_list, length_list, sec_type):
    if len(freq_list) != len(length_list):
        raise KeyboardInterrupt('config_c_pack_20250519_with_tb error: 列表长度不等')

    c_num_list = []
    c_imp_list = []
    c_val_list = []

    for index in range(len(freq_list)):
        freq = freq_list[index]
        length = length_list[index]

        c_val = config_c_value_20250519_with_tb(freq, sec_type)
        c_num = config_c_num_20250519_with_tb(freq, length, sec_type)

        val_tmp = c_val * 1e-6
        c_imp = ImpedanceMultiFreq()
        c_imp.rlc_s = {
            1700: [10e-3, None, val_tmp],
            2000: [10e-3, None, val_tmp],
            2300: [10e-3, None, val_tmp],
            2600: [10e-3, None, val_tmp]}

        c_num_list.append(c_num)
        c_val_list.append(c_val)
        c_imp_list.append(c_imp)

    ret = {
        '电容数量列表': c_num_list,
        '电容容值列表': c_val_list,
        '电容阻抗列表': c_imp_list,
    }

    return ret


# 配置行数据
def config_row_data_20250519_with_tb(df_input, para, data):
    # 序号
    data['序号'] = para['序号'] = df_input['序号']

    # 备注
    data['备注'] = para['备注'] = '无'

    # 区段名
    data['主串区段'] = para['主串区段'] = ''
    data['被串区段'] = para['被串区段'] = ''

    # 区段类型
    data['主串区段类型'] = df_input['主串类型']
    data['被串区段类型'] = df_input['被串类型']

    # 区段长度
    length1 = data['主串区段长度(m)'] = df_input['区段长度']
    length2 = data['被串区段长度(m)'] = df_input['区段长度']
    para['主串区段长度'] = [length1]
    para['被串区段长度'] = [length2, length2, length2]

    # 相对位置
    data['被串相对位置(m)'] = offset = df_input['相对位置']

    para['offset_zhu'] = 0
    para['offset_bei'] = offset - length2

    # 耦合系数
    data['耦合系数(μH/km)'] = para['耦合系数'] = 20

    # 区段频率
    para['freq_主'] = freq1 = data['主串频率(Hz)'] = df_input['主串频率']
    para['freq_被'] = freq2 = data['被串频率(Hz)'] = df_input['被串频率']
    freq = freq1
    data['freq'] = para['freq'] = Freq(freq1)
    para['主串频率列表'] = [Freq(freq1)]
    para['被串频率列表'] = generate_frqs(Freq(freq2), 3, flip_flag=True)

    # 电容配置
    c_pack_zhu = config_c_pack_20250519_with_tb(para['主串频率列表'], para['主串区段长度'], data['主串区段类型'])
    c_pack_bei = config_c_pack_20250519_with_tb(para['被串频率列表'], para['被串区段长度'], data['被串区段类型'])

    # 电容数量
    para['主串电容数'] = c_pack_zhu['电容数量列表']
    para['被串电容数'] = c_pack_bei['电容数量列表']

    data['主串电容数(含TB)'] = para['主串电容数'][0]
    data['被串电容数(含TB)'] = para['被串电容数'][1]

    # 电容容值
    data['主串电容值(μF)'] = c_pack_zhu['电容容值列表'][0]
    data['被串电容值(μF)'] = c_pack_bei['电容容值列表'][1]

    data['主串电容数量列表'] = c_pack_zhu['电容数量列表']
    data['被串电容数量列表'] = c_pack_bei['电容数量列表']

    data['主串电容容值列表'] = c_pack_zhu['电容容值列表']
    data['被串电容容值列表'] = c_pack_bei['电容容值列表']

    para['主串容值列表'] = c_pack_zhu['电容阻抗列表']
    para['被串容值列表'] = c_pack_bei['电容阻抗列表']

    # 道床电阻
    rd = 10000
    data['主串道床电阻(Ω·km)'] = rd
    data['被串道床电阻(Ω·km)'] = rd

    para['主串道床电阻'] = Constant(data['主串道床电阻(Ω·km)'])
    para['被串道床电阻'] = Constant(data['被串道床电阻(Ω·km)'])

    para['Rd'].value = rd

    # 钢轨阻抗
    data['钢轨电阻(Ω/km)'] = round(para['Trk_z'].rlc_s[freq][0], 10)
    data['钢轨电感(H/km)'] = round(para['Trk_z'].rlc_s[freq][1], 10)

    para['主串钢轨阻抗'] = para['Trk_z']
    para['被串钢轨阻抗'] = para['Trk_z']

    # 发码方向
    data['主串方向'] = para['sr_mod_主'] = '右发'
    data['被串方向'] = para['sr_mod_被'] = '右发'

    # 电缆参数
    data['电缆电阻最大(Ω/km)'] = 45
    data['电缆电阻最小(Ω/km)'] = 43
    data['电缆电容最大(F/km)'] = 28e-9
    data['电缆电容最小(F/km)'] = 28e-9

    para['Cable_R'].value = data['电缆电阻最小(Ω/km)']
    para['Cable_C'].value = data['电缆电容最大(F/km)']

    # 电缆长度
    cab_len = 10
    para['cab_len'] = cab_len
    data['主串电缆长度(km)'] = para['主串电缆长度'] = cab_len
    data['被串电缆长度(km)'] = para['被串电缆长度'] = cab_len

    # 分路电阻

    sht_position = df_input['主串分路位置']

    r_sht = 1e-7

    if sht_position == '调整':
        data['主串分路电阻(Ω)'] = para['主串分路电阻'] = 1e10
        data['被串分路电阻(Ω)'] = para['被串分路电阻'] = r_sht
        data['主串分路电阻(Ω)'] = 'None'
        para['主串分路位置'] = 0
        data['主串分路位置(m)'] = '主串调整'
    else:
        data['主串分路电阻(Ω)'] = para['主串分路电阻'] = r_sht
        data['被串分路电阻(Ω)'] = para['被串分路电阻'] = r_sht

        if sht_position == '同位置':
            para['主串分路位置'] = '同位置'
            data['主串分路位置(m)'] = '与被串位置相同'
        else:
            para['主串分路位置'] = length2 - 14.5 - sht_position
            data['主串分路位置(m)'] = sht_position

    para['Rsht_z'] = r_sht

    # 功出电源
    data['主串电平级'] = para['send_level'] = 3
    data['电源电压'] = para['pwr_v_flg'] = '最大'

    # 特殊位置
    data['极性交叉位置'] = para['极性交叉位置'] = []
    data['特殊位置'] = para['special_point'] = data['极性交叉位置']
    data['节点选取模式'] = para['节点选取模式'] = '特殊'

    # 机车信号
    data['最小机车信号位置'] = '-'
    data['机车信号感应系数'] = \
        str(para['机车信号比例V']) + '/' + str(para['机车信号比例I'][freq])
    para['机车信号系数值'] = para['机车信号比例V'] / para['机车信号比例I'][freq]

    # 分路间隔
    data['分路间隔(m)'] = 1
    data['分路起点'] = offset - 14.5
    data['分路终点'] = offset + length2 + 14.5

    # 换装TB
    flag = df_input['是否换装TB']
    if flag:
        data['是否换装TB'] = '是'
        para['section_type'] = '2000A_QJ_With_TB'
    else:
        data['是否换装TB'] = '否'
        para['section_type'] = '2000A'


class PreModel_QJ_20250519_with_tb(PreModel):
    def __init__(self, parameter):
        # super().__init__(turnout_list, parameter)
        self.parameter = para = parameter
        self.train1 = Train(name_base='列车1', posi=0, parameter=parameter)
        self.train2 = Train(name_base='列车2', posi=0, parameter=parameter)
        self.train1['分路电阻1'].z = para['被串分路电阻']
        self.train2['分路电阻1'].z = para['主串分路电阻']

        # 轨道电路初始化
        send_level = para['send_level']

        sg3 = SectionGroup(name_base='地面', posi=para['offset_zhu'], m_num=1,
                           m_frqs=para['主串频率列表'],
                           m_lens=para['主串区段长度'],
                           j_lens=[29, 29],
                           m_typs=[para['section_type']],
                           c_nums=para['主串电容数'],
                           sr_mods=[para['sr_mod_主']],
                           send_lvs=[send_level],
                           parameter=parameter)

        flg = para['pwr_v_flg']
        if para['sr_mod_主'] == '左发':
            sg3['区段1']['左调谐单元'].set_power_voltage(flg)
        elif para['sr_mod_主'] == '右发':
            sg3['区段1']['右调谐单元'].set_power_voltage(flg)

        freq_tmp = Freq(para['freq_被'])
        freq_tmp.change_freq()

        m_num = len(para['被串区段长度'])
        j_num = m_num + 1
        sg4 = SectionGroup(name_base='地面', posi=para['offset_bei'], m_num=m_num,
                           m_frqs=para['被串频率列表'],
                           m_lens=para['被串区段长度'],
                           j_lens=[29] * j_num,
                           m_typs=[para['section_type']] * m_num,
                           c_nums=para['被串电容数'],
                           sr_mods=[para['sr_mod_被']] * m_num,
                           send_lvs=[send_level] * m_num,
                           parameter=parameter)

        self.section_group3 = sg3
        self.section_group4 = sg4

        self.change_c_value()

        self.l3 = l3 = Line(name_base='线路3', sec_group=sg3,
                            parameter=parameter)
        self.l4 = l4 = Line(name_base='线路4', sec_group=sg4,
                            parameter=parameter)
        self.set_rail_para(line=l3, z_trk=para['Trk_z'], rd=para['Trk_z'])
        self.set_rail_para(line=l4, z_trk=para['Trk_z'], rd=para['Trk_z'])

        self.lg = LineGroup(l3, l4, name_base='线路组')

        self.lg.special_point = para['special_point']
        self.lg.refresh()

    def change_c_value(self):
        para = self.parameter

        for index, sec in enumerate(self.section_group3.element.values()):
            for ele in sec.element.values():
                if isinstance(ele, CapC):
                    ele.z = para['主串容值列表'][index]

        for index, sec in enumerate(self.section_group4.element.values()):
            for ele in sec.element.values():
                if isinstance(ele, CapC):
                    ele.z = para['被串容值列表'][index]

    def add_train(self):
        para = self.parameter
        l3 = Line(name_base='线路3', sec_group=self.section_group3,
                  parameter=self.parameter, train=[self.train2])
        self.l3 = l3

        l4 = Line(name_base='线路4', sec_group=self.section_group4,
                  parameter=self.parameter, train=[self.train1])
        self.l4 = l4

        self.set_rail_para(line=l3, z_trk=para['主串钢轨阻抗'], rd=para['主串道床电阻'])
        self.set_rail_para(line=l4, z_trk=para['被串钢轨阻抗'], rd=para['被串道床电阻'])

        self.lg = LineGroup(self.l3, self.l4, name_base='线路组')
        self.lg.special_point = self.parameter['special_point']
        self.lg.refresh()


def write_to_excel2(df, writer, sheet_name, format_dict=None):
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    if format_dict is not None:
        workbook = writer.book
        header_format = workbook.add_format(format_dict)

        worksheet = writer.sheets[sheet_name]
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)


def generate_data_df():
    import os
    import time

    timestamp = time.strftime("%Y%m%d%H%M%S", time.localtime())

    # dir_path = 'C:\\Users\\李继隆\\Desktop\\区间轨道电路换装TB\\结果数据'
    dir_path = 'C:\\Users\\李继隆\\Desktop\\区间轨道电路换装TB\\同上下行\\结果数据'
    new_dir = '%s\\数据简化' % dir_path

    if not os.path.exists(new_dir):
        os.makedirs(new_dir)

    ret = pd.DataFrame()
    sheet_name = '数据输出'

    for file in os.listdir(dir_path):
        if file[-5:] != '.xlsx':
            continue

        file_path = os.path.join(dir_path, file)
        df_input = pd.read_excel(file_path, sheet_name)
        print(df_input)
        ret = pd.concat([ret, df_input])

    new_path = '%s\\区间更换TB_数据简化_%s.xlsx' % (new_dir, timestamp)

    writer = pd.ExcelWriter(new_path, engine='xlsxwriter')
    format_dict = {
        'bold': True,  # 字体加粗
        'text_wrap': True,  # 是否自动换行
        'valign': 'vcenter',  # 垂直对齐方式
        'align': 'center',  # 水平对齐方式
        'border': 1
    }

    write_to_excel2(ret, writer, sheet_name, format_dict=format_dict)
    writer.save()


def generate_data_df2(sub_str):
    import os
    # import time

    # timestamp = time.strftime("%Y%m%d%H%M%S", time.localtime())

    dir_path = 'C:\\Users\\李继隆\\Desktop\\区间轨道电路换装TB\\同上下行\\结果数据'
    new_dir = '%s\\数据简化' % dir_path

    if not os.path.exists(new_dir):
        os.makedirs(new_dir)

    ret = pd.DataFrame()

    for file in os.listdir(dir_path):
        if file[-5:] != '.xlsx':
            continue
        if sub_str != file.split('_')[2]:
            continue

        file_path = os.path.join(dir_path, file)
        df1 = pd.read_excel(file_path, '数据输出')
        df2 = pd.read_excel(file_path, '被串钢轨电流')
        df = pd.concat([df1, df2], axis=1)

        df = df[df['被串相对位置(m)'] == 0]
        print(df.shape)
        ret = pd.concat([ret, df])

    ret['序号'] = list(range(1, ret.shape[0] + 1))
    ret = ret.set_index('序号', drop=False)
    print(ret)

    new_path = '%s\\区间更换TB_数据简化_%s.xlsx' % (new_dir, sub_str)

    writer = pd.ExcelWriter(new_path, engine='xlsxwriter')
    format_dict = {
        'bold': True,  # 字体加粗
        'text_wrap': True,  # 是否自动换行
        'valign': 'vcenter',  # 垂直对齐方式
        'align': 'center',  # 水平对齐方式
        'border': 1
    }

    write_to_excel2(ret, writer, '被串钢轨电流', format_dict=format_dict)
    writer.save()


def math_rm(src: str):
    # src = r'门限值75\%'
    pattern = r'[a-zA-Z0-9()\\\%\-]+'

    ret = list()

    start = 0
    for val in re.finditer(pattern, src):
        ret.append(src[start: val.start()])
        start = val.end()
        ret.append(r'$\mathrm{%s}$' % val.group())
    ret.append(src[start:])

    ret = r''.join(ret)
    return ret


def draw_image_20250526_with_tb():
    # plt.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签
    # plt.rcParams['font.sans-serif'] = ['SimSun']  # 用来正常显示中文标签
    # plt.rcParams['font.sans-serif'] = ['STSong']  # 用来正常显示中文标签
    # plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']  # 用来正常显示中文标签
    # plt.rcParams['font.sans-serif'] = ['consolas']  # 用来正常显示中文标签
    # plt.rcParams['axes.unicode_minus'] = False

    plt.rcParams['font.sans-serif'] = ['SimSun']
    plt.rcParams['mathtext.fontset'] = 'stix'
    plt.rcParams['axes.unicode_minus'] = False

    timestamp = time.strftime("%Y%m%d%H%M%S", time.localtime())
    save_dir = 'C:\\Users\\李继隆\\Desktop\\区间轨道电路换装TB\\同上下行\\图表汇总\\图表汇总_%s' % timestamp

    length_list = [
        600,
        1000,
        1200,
        1400,

    ]

    for sec_length in length_list:
        root = 'C:\\Users\\李继隆\\Desktop\\区间轨道电路换装TB\\同上下行\\结果数据\\数据简化\\'
        file = '区间更换TB_数据简化_%sm.xlsx' % sec_length
        df_input = pd.read_excel(root + file, '被串钢轨电流')

        # freq_list1 = [
        #     [1700, 2000],
        #     [1700, 2600],
        #     [2300, 2000],
        #     [2300, 2600],
        # ]
        #
        # freq_list2 = [
        #     [2000, 1700],
        #     [2000, 2300],
        #     [2600, 1700],
        #     [2600, 2300],
        # ]

        freq_list1 = [
            [1700, 1700],
            [1700, 2300],
            [2300, 1700],
            [2300, 2300],
        ]

        freq_list2 = [
            [2000, 2000],
            [2000, 2600],
            [2600, 2000],
            [2600, 2600],
        ]

        list0 = [freq_list1, freq_list2]

        for i, freq_list in enumerate(list0):

            # 创建图表
            fig = plt.figure(figsize=(16, 8), dpi=100)
            # fig.subplots_adjust(hspace=0.4, wspace=0.1, top=0.8, left=0.15, right=0.85)
            fig.subplots_adjust(hspace=0.3, wspace=0.1, top=0.87, left=0.15, right=0.85)
            # fig.subplots_adjust(hspace=0.4)
            title = math_rm('区段长度%sm' % sec_length)
            fig.suptitle(title, x=0.5, y=0.98, fontsize=25, fontfamily='SimHei')

            ax_list = []

            for j, val in enumerate(freq_list):
                freq1 = val[0]
                freq2 = val[1]

                cnd = (df_input['主串频率(Hz)'] == freq1) & (df_input['被串频率(Hz)'] == freq2)
                df1 = df_input.loc[cnd].copy()
                columns = [col for col in df1.columns if isinstance(col, (int, float))]

                df1 = df1[columns]
                df1 = df1[df1.columns] * 1000
                print(df1)

                cnd = (df_input['主串分路位置(m)'] == '主串调整') & (df_input['是否换装TB'] == '是')
                s1 = df1.loc[cnd].iloc[0, :].copy()

                cnd = (df_input['主串分路位置(m)'] == '主串调整') & (df_input['是否换装TB'] == '否')
                s2 = df1.loc[cnd].iloc[0, :].copy()

                cnd = (df_input['主串分路位置(m)'] == '与被串位置相同') & (df_input['是否换装TB'] == '是')
                s3 = df1.loc[cnd].iloc[0, :].copy()

                cnd = (df_input['主串分路位置(m)'] == '与被串位置相同') & (df_input['是否换装TB'] == '否')
                s4 = df1.loc[cnd].iloc[0, :].copy()

                # 添加子图
                ax = fig.add_subplot(2, 2, j + 1)
                ax_list.append(ax)

                sub_title = math_rm('主串频率%sHz-被串频率%sHz' % (freq1, freq2))
                ax.set_title(sub_title, pad=8, fontsize=12)

                # 纵坐标
                # # ax.yaxis.grid(True, which='major')
                # y_ticks = [0, 100, 200, 300]
                # y_label = map(lambda x: '%.0f' % x, y_ticks)
                #
                # ax.set_yticks(y_ticks)
                # ax.set_yticklabels(y_label)
                #
                # # ax.yaxis.set_font(20)
                # ax.set_ylim([0, 350])

                # 横坐标
                t = np.arange(100, sec_length, 100)
                x_ticks = [0, 29, sec_length, sec_length + 29] + list(t + 14.5)
                x_label = ['发\n送', '接\n收', '发\n送', '接\n收'] + list(t)
                x_label = list(map(lambda x: math_rm(str(x)), x_label))

                ax.set_xticks(x_ticks)
                ax.set_xticklabels(x_label)

                # if pos_index in [13, 14, 15, 16]:
                #     ax.set_xticklabels(x_label)
                # else:
                #     ax.set_xticklabels([''] * len(x_ticks))
                # ax.set_yticklabels(fontfamily="Times New Roman")

                # 坐标轴字体
                ax.tick_params(
                    # axis='y',
                    labelsize=9,  # y轴字体大小设置
                    # color='r',        # y轴标签颜色设置
                    # labelcolor='b',   # y轴字体颜色设置
                    direction='in',  # y轴标签方向设置
                    # pad=10,
                )

                ###################################

                xx1 = range(len(s1))

                ax.plot(xx1, s2.values, linestyle='-', alpha=1, color='blue', label=math_rm('主串调整'))
                ax.plot(xx1, s1.values, linestyle='--', alpha=0.5, color='blue', label=math_rm('主串调整-换装TB'))
                ax.plot(xx1, s4.values, linestyle='-', alpha=1, color='green', label=math_rm('主串同位置分路'))
                ax.plot(xx1, s3.values, linestyle='--', alpha=0.5, color='green', label=math_rm('主串同位置分路-换装TB'))

                threshold_dict = {
                    1700: 263,
                    2000: 234,
                    2300: 217,
                    2600: 200,
                }

                threshold = threshold_dict[freq1]

                length_x = sec_length
                xx = np.arange(length_x)
                # yy = np.ones(length_x) * min(max_list)

                yy2 = np.ones(length_x) * threshold
                yy3 = yy2 * 0.75

                # ax.plot(xx, yy, linestyle='--', alpha=0.8, color='blue', label='最优值')
                ax.plot(xx, yy2, linestyle='--', alpha=0.8, color='orange', label=math_rm(r'门限值'))
                ax.plot(xx, yy3, linestyle='--', alpha=0.8, color='r', label=math_rm(r'门限值75\%'))

                pos_x = length_x + 10
                ax.annotate(math_rm(r'%.0fmA' % yy2[0]), (pos_x, yy2[0]), xytext=(pos_x + 30, yy2[0] + 10), ha="right",
                            fontsize=9, color='orange')
                ax.annotate(math_rm(r'%.0fmA' % yy3[0]), (pos_x, yy3[0]), xytext=(pos_x + 30, yy3[0] + 10), ha="right",
                            fontsize=9, color='r')

                # # y轴 双坐标轴
                # ax2 = ax.twinx()
                #
                # ax2.set_ylim([0, 300])
                #
                # y2_ticks = [threshold, threshold * 0.75]
                # y2_label = map(lambda x: r'$\mathrm{%.0fmA}$' % x, y2_ticks)
                #
                # ax2.set_yticks(y2_ticks)
                # ax2.set_yticklabels(y2_label)
                # ax2.tick_params(labelsize=9, direction='in')
                #
                # ax2.get_yticklabels()[0].set_color('orange')
                # ax2.get_yticklabels()[1].set_color('r')

                # # 箭头
                # s_tmp = df_data1.iloc[max_index, :]
                # arrow_x = s_tmp["被串最大干扰位置(m)"]
                # arrow_y = s_tmp["被串最大干扰电流(A)"] * 1000
                # offset_tmp = s_tmp["被串相对位置(m)"]
                #
                # txt = '最大干扰电流$\mathrm{%.2fmA}$\n主被串错位$\mathrm{%.0fm}$' % (arrow_y, offset_tmp)
                #
                # txt_y = 300 if arrow_y > 300 else arrow_y
                # txt_x = 50 if arrow_x < 50 else arrow_x
                # ax.annotate(
                #     txt, (arrow_x, arrow_y),
                #     xytext=(txt_x + 100, txt_y + 50),
                #     ha="center", va="center",
                #     # textcoords='offset points',
                #     fontsize=10,
                #     color='blue',
                #     arrowprops=dict(
                #         # facecolor='#74C476',
                #         alpha=0.6,
                #         arrowstyle='fancy',
                #         # connectionstyle='arc3,rad=0.5',
                #         color='blue',
                #     )
                # )

            plt.text(
                0.5, 0.07, math_rm(r'被串分路位置(m)'),
                va='top', ha='center', transform=fig.transFigure,
                fontsize=13,
            )

            plt.text(
                0.12, 0.5, math_rm(r'邻线干扰电流(mA)'),
                va='center', ha='right', transform=fig.transFigure,
                fontsize=13, rotation=90,
            )

            handles, labels = ax_list[0].get_legend_handles_labels()
            plt.legend(
                handles, labels,
                loc='center right',
                # ncol=3,
                bbox_to_anchor=(1.44, 1.2),
                fontsize=11,
            )

            # plt.show()
            # raise KeyboardInterrupt()

            if not os.path.exists(save_dir):
                os.makedirs(save_dir)

            filename1 = '%s\\区间更换TB邻线干扰_%sm_%s.png' % (save_dir, sec_length, i+1)
            MainLog.add_log_accurate('save figure --> %s' % filename1)
            fig.savefig(filename1, transparent=True)


def generate_data_df3(sec_length):
    import os
    # import time

    # timestamp = time.strftime("%Y%m%d%H%M%S", time.localtime())

    dir_path = 'C:\\Users\\李继隆\\Desktop\\区间轨道电路换装TB\\同上下行\\结果数据\\数据简化'
    # file_path = os.path.join(dir_path, '区间更换TB_数据简化_20250526114533.xlsx')
    file_path = os.path.join(dir_path, '区间更换TB_数据简化_20250811092706.xlsx')

    df = pd.read_excel(file_path, '整理2')

    df1 = df[df['主串区段长度(m)'] == sec_length].copy()

    new_path = '%s\\区间更换TB_数据整理_%sm.xlsx' % (dir_path, sec_length)

    writer = pd.ExcelWriter(new_path, engine='xlsxwriter')
    format_dict = {
        'bold': True,  # 字体加粗
        'text_wrap': True,  # 是否自动换行
        'valign': 'vcenter',  # 垂直对齐方式
        'align': 'center',  # 水平对齐方式
        'border': 1
    }

    write_to_excel2(df1, writer, '数据整理', format_dict=format_dict)
    writer.save()


def draw_image_20250526_with_tb2():
    # plt.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签
    # plt.rcParams['font.sans-serif'] = ['SimSun']  # 用来正常显示中文标签
    # plt.rcParams['font.sans-serif'] = ['STSong']  # 用来正常显示中文标签
    # plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']  # 用来正常显示中文标签
    # plt.rcParams['font.sans-serif'] = ['consolas']  # 用来正常显示中文标签
    # plt.rcParams['axes.unicode_minus'] = False

    plt.rcParams['font.sans-serif'] = ['SimSun']
    plt.rcParams['mathtext.fontset'] = 'stix'
    plt.rcParams['axes.unicode_minus'] = False

    timestamp = time.strftime("%Y%m%d%H%M%S", time.localtime())
    save_dir = 'C:\\Users\\李继隆\\Desktop\\区间轨道电路换装TB\\同上下行\\图表汇总\\错位遍历_%s' % timestamp

    if not os.path.exists(save_dir):
        os.makedirs(save_dir)

    length_list = [
        600,
        1000,
        1200,
        1400,

    ]

    for sec_length in length_list:
        root = 'C:\\Users\\李继隆\\Desktop\\区间轨道电路换装TB\\同上下行\\结果数据\\数据简化\\'
        file = '区间更换TB_数据整理_%sm.xlsx' % sec_length
        df_input = pd.read_excel(root + file, '数据整理')

        # freq_list1 = [
        #     [1700, 2000],
        #     [1700, 2600],
        #     [2300, 2000],
        #     [2300, 2600],
        # ]
        #
        # freq_list2 = [
        #     [2000, 1700],
        #     [2000, 2300],
        #     [2600, 1700],
        #     [2600, 2300],
        # ]

        freq_list1 = [
            [1700, 1700],
            [1700, 2300],
            [2300, 1700],
            [2300, 2300],
        ]

        freq_list2 = [
            [2000, 2000],
            [2000, 2600],
            [2600, 2000],
            [2600, 2600],
        ]

        list0 = [freq_list1, freq_list2]

        for i, freq_list in enumerate(list0):

            # 创建图表
            fig = plt.figure(figsize=(16, 8), dpi=100)
            # fig.subplots_adjust(hspace=0.4, wspace=0.1, top=0.8, left=0.15, right=0.85)
            fig.subplots_adjust(hspace=0.3, wspace=0.1, top=0.87, left=0.15, right=0.85)
            # fig.subplots_adjust(hspace=0.4)
            title = math_rm('区段长度%sm' % sec_length)
            fig.suptitle(title, x=0.5, y=0.98, fontsize=25, fontfamily='SimHei')

            ax_list = []

            for j, val in enumerate(freq_list):
                freq1 = val[0]
                freq2 = val[1]

                cnd = (df_input['主串频率(Hz)'] == freq1) & (df_input['被串频率(Hz)'] == freq2)
                df1 = df_input.loc[cnd].copy()
                print(df1)

                offset = df1.loc[:, '被串相对位置(m)'].copy()
                s1 = df1.loc[:, '被串最大干扰电流(A)'].copy() / 0.001
                s2 = df1.loc[:, '换TB被串最大干扰电流(A)'].copy() / 0.001
                s3 = df1.loc[:, '同位置-被串最大干扰电流(A)'].copy() / 0.001
                s4 = df1.loc[:, '同位置-换TB被串最大干扰电流(A)'].copy() / 0.001

                # 添加子图
                ax = fig.add_subplot(2, 2, j + 1)
                ax_list.append(ax)

                sub_title = math_rm('主串频率%sHz-被串频率%sHz' % (freq1, freq2))
                ax.set_title(sub_title, pad=8, fontsize=12)

                # 纵坐标
                # # ax.yaxis.grid(True, which='major')
                # y_ticks = [0, 100, 200, 300]
                # y_label = map(lambda x: '%.0f' % x, y_ticks)
                #
                # ax.set_yticks(y_ticks)
                # ax.set_yticklabels(y_label)
                #
                # # ax.yaxis.set_font(20)
                # ax.set_ylim([0, 350])

                # 横坐标
                x_ticks = list(range(len(offset)))

                x_label = list(offset)
                x_label = list(map(lambda x: math_rm(str(x)), x_label))

                print(x_ticks)
                print(x_label)

                ax.set_xticks(x_ticks)
                ax.set_xticklabels(x_label)

                # if pos_index in [13, 14, 15, 16]:
                #     ax.set_xticklabels(x_label)
                # else:
                #     ax.set_xticklabels([''] * len(x_ticks))
                # ax.set_yticklabels(fontfamily="Times New Roman")

                # 坐标轴字体
                ax.tick_params(
                    # axis='y',
                    labelsize=9,  # y轴字体大小设置
                    # color='r',        # y轴标签颜色设置
                    # labelcolor='b',   # y轴字体颜色设置
                    direction='in',  # y轴标签方向设置
                    # pad=10,
                )

                ###################################

                xx1 = range(len(s1))

                # ax.plot(xx1, s1.values, linestyle='--', alpha=0.5, color='blue', label=math_rm('主串调整'))
                # ax.plot(xx1, s2.values, linestyle='-', alpha=1, color='blue', label=math_rm('主串调整-换装TB'))
                ax.plot(xx1, s3.values, linestyle='--', alpha=0.5, color='green', label=math_rm('主串同位置分路'))
                ax.plot(xx1, s4.values, linestyle='-', alpha=1, color='green', label=math_rm('主串同位置分路-换装TB'))

                threshold_dict = {
                    1700: 263,
                    2000: 234,
                    2300: 217,
                    2600: 200,
                }

                threshold = threshold_dict[freq1]

                length_x = len(xx1)
                xx = np.arange(length_x)
                # yy = np.ones(length_x) * min(max_list)

                yy2 = np.ones(length_x) * threshold
                yy3 = yy2 * 0.75

                # ax.plot(xx, yy, linestyle='--', alpha=0.8, color='blue', label='最优值')
                ax.plot(xx, yy2, linestyle='--', alpha=0.8, color='orange', label=math_rm(r'门限值'))
                ax.plot(xx, yy3, linestyle='--', alpha=0.8, color='r', label=math_rm(r'门限值75\%'))

                pos_x = length_x-1
                ax.annotate(math_rm(r'%.0fmA' % yy2[0]), (pos_x, yy2[0]), xytext=(pos_x, yy2[0]+10), ha="right",
                            fontsize=9, color='orange')
                ax.annotate(math_rm(r'%.0fmA' % yy3[0]), (pos_x, yy3[0]), xytext=(pos_x, yy3[0]+10), ha="right",
                            fontsize=9, color='r')

                # # y轴 双坐标轴
                # ax2 = ax.twinx()
                #
                # ax2.set_ylim([0, 300])
                #
                # y2_ticks = [threshold, threshold * 0.75]
                # y2_label = map(lambda x: r'$\mathrm{%.0fmA}$' % x, y2_ticks)
                #
                # ax2.set_yticks(y2_ticks)
                # ax2.set_yticklabels(y2_label)
                # ax2.tick_params(labelsize=9, direction='in')
                #
                # ax2.get_yticklabels()[0].set_color('orange')
                # ax2.get_yticklabels()[1].set_color('r')

                # # 箭头
                # s_tmp = df_data1.iloc[max_index, :]
                # arrow_x = s_tmp["被串最大干扰位置(m)"]
                # arrow_y = s_tmp["被串最大干扰电流(A)"] * 1000
                # offset_tmp = s_tmp["被串相对位置(m)"]
                #
                # txt = '最大干扰电流$\mathrm{%.2fmA}$\n主被串错位$\mathrm{%.0fm}$' % (arrow_y, offset_tmp)
                #
                # txt_y = 300 if arrow_y > 300 else arrow_y
                # txt_x = 50 if arrow_x < 50 else arrow_x
                # ax.annotate(
                #     txt, (arrow_x, arrow_y),
                #     xytext=(txt_x + 100, txt_y + 50),
                #     ha="center", va="center",
                #     # textcoords='offset points',
                #     fontsize=10,
                #     color='blue',
                #     arrowprops=dict(
                #         # facecolor='#74C476',
                #         alpha=0.6,
                #         arrowstyle='fancy',
                #         # connectionstyle='arc3,rad=0.5',
                #         color='blue',
                #     )
                # )

            plt.text(
                0.5, 0.07, math_rm(r'主被串相对位置(m)'),
                va='top', ha='center', transform=fig.transFigure,
                fontsize=13,
            )

            plt.text(
                0.12, 0.5, math_rm(r'最大邻线干扰电流(mA)'),
                va='center', ha='right', transform=fig.transFigure,
                fontsize=13, rotation=90,
            )

            handles, labels = ax_list[0].get_legend_handles_labels()
            plt.legend(
                handles, labels,
                loc='center right',
                # ncol=3,
                bbox_to_anchor=(1.44, 1.2),
                fontsize=11,
            )

            # plt.show()
            # raise KeyboardInterrupt()

            filename1 = '%s\\区间更换TB邻线干扰_错位遍历_%sm_%s.png' % (save_dir, sec_length, i+1)
            MainLog.add_log_accurate('save figure --> %s' % filename1)
            fig.savefig(filename1, transparent=True)


def draw_image_20250526_with_tb3():
    # plt.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签
    # plt.rcParams['font.sans-serif'] = ['SimSun']  # 用来正常显示中文标签
    # plt.rcParams['font.sans-serif'] = ['STSong']  # 用来正常显示中文标签
    # plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']  # 用来正常显示中文标签
    # plt.rcParams['font.sans-serif'] = ['consolas']  # 用来正常显示中文标签
    # plt.rcParams['axes.unicode_minus'] = False

    plt.rcParams['font.sans-serif'] = ['SimSun']
    plt.rcParams['mathtext.fontset'] = 'stix'
    plt.rcParams['axes.unicode_minus'] = False

    timestamp = time.strftime("%Y%m%d%H%M%S", time.localtime())
    save_dir = 'C:\\Users\\李继隆\\Desktop\\区间轨道电路换装TB\\同上下行\\图表汇总\\变化率_%s' % timestamp

    length_list = [
        600,
        1000,
        1200,
        1400,
    ]

    for sec_length in length_list:
        root = 'C:\\Users\\李继隆\\Desktop\\区间轨道电路换装TB\\同上下行\\结果数据\\数据简化\\'
        file = '区间更换TB_数据整理_%sm.xlsx' % sec_length
        df_input = pd.read_excel(root + file, '数据整理')

        # list0 = [
        #     [1700, 2000],
        #     [1700, 2600],
        #     [2300, 2000],
        #     [2300, 2600],
        #     [2000, 1700],
        #     [2000, 2300],
        #     [2600, 1700],
        #     [2600, 2300],
        # ]

        list0 = [
            [1700, 1700],
            [1700, 2300],
            [2300, 1700],
            [2300, 2300],
            [2000, 2000],
            [2000, 2600],
            [2600, 2000],
            [2600, 2600],
        ]
        # list0 = [freq_list1, freq_list2, freq_list3, freq_list4]

        for i, val in enumerate(list0):

            freq1 = val[0]
            freq2 = val[1]

            # 创建图表
            # fig = plt.figure(figsize=(16, 8), dpi=100)
            fig, ax_list = plt.subplots(2, 2, figsize=(16, 6), dpi=100, gridspec_kw={'height_ratios': [3, 1]})

            # fig.subplots_adjust(hspace=0.4, wspace=0.1, top=0.8, left=0.15, right=0.85)
            fig.subplots_adjust(hspace=0.3, wspace=0.1, top=0.87, left=0.15, right=0.85)
            # fig.subplots_adjust(hspace=0.4)
            title = math_rm('区段长度%sm-主串%sHz-被串%sHz' % (sec_length, freq1, freq2))
            fig.suptitle(title, x=0.5, y=0.98, fontsize=25, fontfamily='SimHei')

            cnd = (df_input['主串频率(Hz)'] == freq1) & (df_input['被串频率(Hz)'] == freq2)
            df1 = df_input.loc[cnd].copy()
            print(df1)
            max_y = 0
            min_y = 0

            handles = []
            labels = []

            for j, status in enumerate([0, 1]):

                offset = df1.loc[:, '被串相对位置(m)'].copy()

                if status == 0:
                    s1 = df1.loc[:, '被串最大干扰电流(A)'].copy() / 0.001
                    s2 = df1.loc[:, '换TB被串最大干扰电流(A)'].copy() / 0.001
                    color = 'blue'
                    legend_label1 = math_rm('主串调整')
                    legend_label2 = math_rm('主串调整-换装TB')
                    sub_title = '主串调整状态-被串分路'
                else:
                    s1 = df1.loc[:, '同位置-被串最大干扰电流(A)'].copy() / 0.001
                    s2 = df1.loc[:, '同位置-换TB被串最大干扰电流(A)'].copy() / 0.001
                    color = 'green'
                    legend_label1 = math_rm('主串同位置分路')
                    legend_label2 = math_rm('主串同位置分路-换装TB')
                    sub_title = '主串被串同位置分路'

                s5 = (s2 / s1 - 1) * 100

                max_rate = max(s5)
                min_rate = min(s5)

                max_y = max(max_y, max(s1), max(s2))
                min_y = min(min_y, min(s1), min(s2))

                max_abs = max(abs(max(s5)), abs(min(s5))) * 1.1

                # 添加子图
                # ax = fig.add_subplot(2, 2, j + 1, gridspec_kw={'height_ratios': [2, 1]})
                # ax_list.append(ax)

                ########################################################################################################

                ax = ax_list.flat[j]

                # sub_title = math_rm('主串频率%sHz-被串频率%sHz' % (freq1, freq2))
                ax.set_title(sub_title, pad=8, fontsize=12)

                # 横坐标
                x_ticks = []
                x_label = []

                for index, offset_val in enumerate(offset):
                    if offset_val % 200 == 0:
                        x_ticks.append(index)
                        x_label.append(offset_val)
                x_label = list(map(lambda x: math_rm(str(x)), x_label))

                print(x_ticks)
                print(x_label)

                ax.set_xticks(x_ticks)
                ax.set_xticklabels(x_label)

                if j == 0:
                    ax.set_ylabel(math_rm(r'最大邻线干扰电流(mA)'))

                # if pos_index in [13, 14, 15, 16]:
                #     ax.set_xticklabels(x_label)
                # else:
                #     ax.set_xticklabels([''] * len(x_ticks))
                # ax.set_yticklabels(fontfamily="Times New Roman")

                # 坐标轴字体
                ax.tick_params(
                    # axis='y',
                    labelsize=9,  # y轴字体大小设置
                    # color='r',        # y轴标签颜色设置
                    # labelcolor='b',   # y轴字体颜色设置
                    direction='in',  # y轴标签方向设置
                    # pad=10,
                )

                xx1 = range(len(s1))

                ax.plot(xx1, s1.values, linestyle='-', alpha=1, color=color, label=legend_label1)
                ax.plot(xx1, s2.values, linestyle='--', alpha=0.5, color=color, label=legend_label2)
                # ax.plot(xx1, s3.values, linestyle='--', alpha=0.5, color='green', label=math_rm('主串同位置分路'))
                # ax.plot(xx1, s4.values, linestyle='-', alpha=1, color='green', label=math_rm('主串同位置分路-换装TB'))

                threshold_dict = {
                    1700: 263,
                    2000: 234,
                    2300: 217,
                    2600: 200,
                }

                threshold = threshold_dict[freq1]

                length_x = len(xx1)
                xx = np.arange(length_x)
                # yy = np.ones(length_x) * min(max_list)

                yy2 = np.ones(length_x) * threshold
                yy3 = yy2 * 0.75

                # ax.plot(xx, yy, linestyle='--', alpha=0.8, color='blue', label='最优值')
                ax.plot(xx, yy2, linestyle='--', alpha=0.8, color='orange', label=math_rm(r'门限值'))
                ax.plot(xx, yy3, linestyle='--', alpha=0.8, color='r', label=math_rm(r'门限值75\%'))

                handles = handles + ax.get_legend_handles_labels()[0]
                labels = labels + ax.get_legend_handles_labels()[1]
                if j == 0:
                    handles = handles[:-2]
                    labels = labels[:-2]

                pos_x = length_x-1
                ax.annotate(math_rm(r'%.0fmA' % yy2[0]), (pos_x, yy2[0]), xytext=(pos_x, yy2[0]+10), ha="right",
                            fontsize=9, color='orange')
                ax.annotate(math_rm(r'%.0fmA' % yy3[0]), (pos_x, yy3[0]), xytext=(pos_x, yy3[0]+10), ha="right",
                            fontsize=9, color='r')

                ########################################################################################################

                ax = ax_list.flat[j+2]
                txt1 = '最小%.2f\%%~最大%.2f\%%' % (min_rate, max_rate)

                ax.set_title(math_rm(txt1), pad=8, fontsize=12)

                # 纵坐标
                # # ax.yaxis.grid(True, which='major')
                # y_ticks = list(range(-100, 101, 20))
                # y_label = list(map(lambda x: '%.0f%%' % x, y_ticks))
                #
                # ax.set_yticks(y_ticks)
                # ax.set_yticklabels(y_label)
                #
                # # ax.yaxis.set_font(20)
                ax.set_ylim([-max_abs, max_abs])
                if j == 0:
                    ax.set_ylabel(math_rm(r'变化率(\%)'))

                # 横坐标
                # x_ticks = list(range(len(offset)))
                #
                # x_label = list(offset)
                # x_label = list(map(lambda x: math_rm(str(x)), x_label))
                #
                # print(x_ticks)
                # print(x_label)

                ax.set_xticks(x_ticks)
                ax.set_xticklabels(x_label)

                # if pos_index in [13, 14, 15, 16]:
                #     ax.set_xticklabels(x_label)
                # else:
                #     ax.set_xticklabels([''] * len(x_ticks))
                # ax.set_yticklabels(fontfamily="Times New Roman")

                # 坐标轴字体
                ax.tick_params(
                    # axis='y',
                    labelsize=9,  # y轴字体大小设置
                    # color='r',        # y轴标签颜色设置
                    # labelcolor='b',   # y轴字体颜色设置
                    direction='in',  # y轴标签方向设置
                    # pad=10,
                )

                ###################################

                xx1 = range(len(s1))

                ax.bar(xx1, s5.values, color=['blue' if v >= 0 else 'red' for v in s5.values], edgecolor='black')

                threshold_dict = {
                    1700: 263,
                    2000: 234,
                    2300: 217,
                    2600: 200,
                }

                threshold = threshold_dict[freq1]

                length_x = len(xx1)

                xx = np.arange(length_x)
                yy = np.ones(length_x) * 0

                ax.plot(xx, yy, linestyle='--', alpha=0.8, color='gray', label='零值')

            ax_list[0, 0].set_ylim(min_y - 5, max_y + 5)
            ax_list[0, 1].set_ylim(min_y - 5, max_y + 5)

            plt.text(
                0.5, 0.07, math_rm(r'主被串相对位置(m)'),
                va='top', ha='center', transform=fig.transFigure,
                fontsize=13,
            )

            # plt.text(
            #     0.86, 0.21, txt,
            #     va='top', ha='left', transform=fig.transFigure,
            #     fontsize=11,
            # )

            # handles, labels = ax_list.flat[0].get_legend_handles_labels()
            plt.legend(
                handles, labels,
                loc='center right',
                # ncol=3,
                bbox_to_anchor=(1.45, 3.2),
                fontsize=11,
            )

            # plt.show()
            # raise KeyboardInterrupt()

            if not os.path.exists(save_dir):
                os.makedirs(save_dir)

            filename1 = '%s\\区间更换TB邻线干扰_变化率_%sm_%s.png' % (save_dir, sec_length, i+1)
            MainLog.add_log_accurate('save figure --> %s' % filename1)
            fig.savefig(filename1, transparent=True)


def generate_data_df4(sec_length):
    import os
    # import time

    # timestamp = time.strftime("%Y%m%d%H%M%S", time.localtime())

    dir_path = 'C:\\Users\\李继隆\\Desktop\\区间轨道电路换装TB\\同上下行\\结果数据\\数据简化'
    # file_path = os.path.join(dir_path, '区间更换TB_数据简化_20250526114533.xlsx')
    file_path = os.path.join(dir_path, '区间更换TB_数据简化_20250811092706.xlsx')

    df = pd.read_excel(file_path, '整理2')

    df1 = df[df['主串区段长度(m)'] == sec_length].copy()

    new_path = '%s\\区间更换TB_干扰变化值_%sm.xlsx' % (dir_path, sec_length)

    # list0 = [
    #     [1700, 2000],
    #     [1700, 2600],
    #     [2300, 2000],
    #     [2300, 2600],
    #     [2000, 1700],
    #     [2000, 2300],
    #     [2600, 1700],
    #     [2600, 2300],
    # ]

    list0 = [
        [1700, 1700],
        [1700, 2300],
        [2300, 1700],
        [2300, 2300],
        [2000, 2000],
        [2000, 2600],
        [2600, 2000],
        [2600, 2600],
    ]

    columns = list(map(lambda x: '主串%s\n被串%s' % (x[0], x[1]), list0))
    columns.insert(0, '被串相对位置(m)')
    columns.insert(0, '主串状态')

    ret1 = pd.DataFrame(columns=columns)
    for _, row in df1.iterrows():
        freq1 = row['主串频率(Hz)']
        freq2 = row['被串频率(Hz)']

        value = (row['换TB被串最大干扰电流(A)'] - row['被串最大干扰电流(A)'])*1000
        value = round(value)
        column = '主串%s\n被串%s' % (freq1, freq2)
        index = '主串调整_%sm' % row['被串相对位置(m)']
        ret1.loc[index, column] = value
        ret1.loc[index, '主串状态'] = '主串调整'
        ret1.loc[index, '被串相对位置(m)'] = row['被串相对位置(m)']

    ret_max = pd.DataFrame()
    ret_min = pd.DataFrame()

    tmp = ret1.copy()
    tmp = tmp.loc[:, tmp.columns[2:]].copy()
    tmp = tmp.astype(float)

    max_idx = tmp.idxmax()
    min_idx = tmp.idxmin()

    ret_max = pd.concat([ret_max, max_idx])
    ret_min = pd.concat([ret_min, min_idx])

    ret2 = pd.DataFrame(columns=columns)
    for _, row in df1.iterrows():
        freq1 = row['主串频率(Hz)']
        freq2 = row['被串频率(Hz)']

        value = (row['同位置-换TB被串最大干扰电流(A)'] - row['同位置-被串最大干扰电流(A)'])*1000
        value = round(value)
        column = '主串%s\n被串%s' % (freq1, freq2)
        index = '主串同位置分路_%sm' % row['被串相对位置(m)']
        ret2.loc[index, column] = value
        ret2.loc[index, '主串状态'] = '主串同位置分路'
        ret2.loc[index, '被串相对位置(m)'] = row['被串相对位置(m)']

    tmp = ret2.copy()
    tmp = tmp.loc[:, tmp.columns[2:]].copy()
    tmp = tmp.astype(float)

    max_idx = tmp.idxmax()
    min_idx = tmp.idxmin()

    ret_max = pd.concat([ret_max, max_idx], axis=1)
    ret_min = pd.concat([ret_min, min_idx], axis=1)

    ret = pd.concat([ret1, ret2])

    writer = pd.ExcelWriter(new_path, engine='xlsxwriter')
    format_dict = {
        'bold': True,  # 字体加粗
        'text_wrap': True,  # 是否自动换行
        'valign': 'vcenter',  # 垂直对齐方式
        'align': 'center',  # 水平对齐方式
        'border': 1
    }

    write_to_excel2(ret, writer, '数据整理', format_dict=format_dict)
    writer.save()

    # 加载 Excel 文件
    wb = load_workbook(new_path)
    ws = wb.active

    red = 'FF0000'
    blue = '00B0F0'

    fill = PatternFill(start_color=blue, end_color=blue, fill_type='solid')
    for column in ret_max.index:
        col_num = ret.columns.get_loc(column)
        for index in ret_max.loc[column, :].values:
            idx_num = ret.index.get_loc(index)
            ws.cell(row=idx_num + 2, column=col_num + 1).fill = fill

    fill = PatternFill(start_color=red, end_color=red, fill_type='solid')
    for column in ret_min.index:
        col_num = ret.columns.get_loc(column)
        for index in ret_min.loc[column, :].values:
            idx_num = ret.index.get_loc(index)
            ws.cell(row=idx_num + 2, column=col_num + 1).fill = fill

    # 插入一行到第 1 行（所有现有内容下移）
    ws.insert_rows(idx=1)

    # 合并新插入行的部分单元格（A1:C1）
    ws.merge_cells('C1:J1')

    # 为合并单元格设置内容和居中对齐
    cell = ws['C1']
    cell.value = '邻线干扰电流变化值(mA)'
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.font = Font(bold=True)

    num = ret.shape[0] / 2

    ws.merge_cells(
        start_row=3,
        start_column=1,
        end_row=num+2,
        end_column=1,
    )
    ws.cell(row=3, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.merge_cells(
        start_row=num+3,
        start_column=1,
        end_row=2*num+2,
        end_column=1,
    )
    ws.cell(row=num+3, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 快捷添加所有框线（数据区域：包括标题行）
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    # 遍历数据区域（从第 2 行到最后一行，所有列）
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border

    # 保存修改后的 Excel 文件
    output_file = '%s\\区间更换TB_干扰变化值(填充颜色)_%sm.xlsx' % (dir_path, sec_length)
    wb.save(output_file)
    print(num)


if __name__ == '__main__':
    # generate_data_df()
    # generate_data_df2('600m')
    # generate_data_df2('1000m')
    # generate_data_df2('1200m')
    # generate_data_df2('1400m')
    # draw_image_20250526_with_tb()
    # generate_data_df3(600)
    # generate_data_df3(1000)
    # generate_data_df3(1200)
    # generate_data_df3(1400)
    # draw_image_20250526_with_tb3()
    generate_data_df4(600)
    generate_data_df4(1000)
    generate_data_df4(1200)
    generate_data_df4(1400)
    pass
